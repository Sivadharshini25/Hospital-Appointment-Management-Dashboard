from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, random
from datetime import date, timedelta, datetime

app = Flask(__name__)
EXCEL_FILE = "hospital_data.xlsx"
random.seed(42)

DOCTORS = [
    {"name": "Dr. Emily White",  "dept": "Cardiology"},
    {"name": "Dr. James Brown",  "dept": "Orthopedics"},
    {"name": "Dr. Jane Doe",     "dept": "Neurology"},
    {"name": "Dr. John Smith",   "dept": "General Medicine"},
    {"name": "Dr. Liam Lee",     "dept": "Pediatrics"},
]

PATIENT_TYPES = ["New Patient", "Follow-up", "Chronic Case", "Emergency", "Consultation"]

RECOMMENDED_SLOTS = {
    "New Patient":  30,
    "Follow-up":    15,
    "Chronic Case": 45,
    "Emergency":    60,
    "Consultation": 20,
}

TIME_SLOTS = [
    "09:00", "09:15", "09:30", "09:45",
    "10:00", "10:15", "10:30", "10:45",
    "11:00", "11:15", "11:30", "11:45",
    "12:00", "12:15", "12:30", "12:45",
    "14:00", "14:15", "14:30", "14:45",
    "15:00", "15:15", "15:30", "15:45",
    "16:00", "16:15", "16:30", "16:45",
]

STATUSES = ["Confirmed", "Completed", "Cancelled", "No-Show"]

HEADERS = [
    "Appointment_ID", "Date", "Patient_Name", "Doctor", "Department",
    "Patient_Type", "Time_Slot", "Allocated_Min", "Actual_Min",
    "Delay_Min", "Status", "Notes"
]

# ── style helpers ────────────────────────────────────────────────────────────
def _thin(): return Side(style="thin", color="D1D5DB")
def _borders(): return Border(left=_thin(), right=_thin(), top=_thin(), bottom=_thin())

def _hdr(cell, bg="2D4A2D"):
    cell.font      = Font(bold=True, color="FFFFFF", name="Times New Roman", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _borders()

def _data(cell, row_i, delay=None):
    cell.font      = Font(name="Times New Roman", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill      = PatternFill("solid", fgColor="F5F8F0" if row_i % 2 == 0 else "FFFFFF")
    cell.border    = _borders()
    if delay is not None:
        if delay > 10:
            cell.fill = PatternFill("solid", fgColor="FEF2F2")
            cell.font = Font(name="Times New Roman", size=9, color="B91C1C", bold=True)
        elif delay < 0:
            cell.fill = PatternFill("solid", fgColor="F0FDF4")
            cell.font = Font(name="Times New Roman", size=9, color="15803D")

# ── seed 100 rows ─────────────────────────────────────────────────────────────
def _seed(ws):
    names = ["Arjun Kumar","Priya Sharma","Ravi Nair","Anita Patel","Vikram Singh",
             "Sunita Das","Mohan Rao","Deepa Iyer","Kiran Mehta","Lakshmi Pillai",
             "Raj Gupta","Meena Joshi","Arun Verma","Kavitha Reddy","Suresh Iyengar",
             "Pooja Mishra","Ganesh Krishnan","Radha Balasubramaniam","Nikhil Shah","Uma Chandrasekaran"]
    today = date.today()
    apt_weights = [20, 35, 15, 10, 20]

    for i in range(100):
        d       = today - timedelta(days=random.randint(0, 44))
        doc     = random.choice(DOCTORS)
        ptype   = random.choices(PATIENT_TYPES, weights=apt_weights)[0]
        slot    = random.choice(TIME_SLOTS)
        alloc   = 15
        rec     = RECOMMENDED_SLOTS[ptype]
        actual  = max(5, int(random.gauss(rec, 6)))
        delay   = actual - alloc
        status  = random.choices(STATUSES, weights=[10, 60, 15, 15])[0]
        name    = random.choice(names)
        apt_id  = f"APT{1000+i}"

        row = [apt_id, str(d), name, doc["name"], doc["dept"],
               ptype, slot, alloc, actual, delay, status, ""]
        ri = i + 2
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            _data(cell, ri, delay=(val if ci == 10 else None))

# ── slot guide sheet ──────────────────────────────────────────────────────────
def _build_slot_guide(wb):
    ws = wb.create_sheet("Slot_Guide")
    ws.sheet_properties.tabColor = "4A7C59"

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = "Recommended Appointment Slot Duration by Patient Type"
    t.font      = Font(bold=True, size=13, color="FFFFFF", name="Times New Roman")
    t.fill      = PatternFill("solid", fgColor="2D4A2D")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value     = "Current system allocates 15 min to all patients. This table shows what each type actually requires."
    s.font      = Font(italic=True, size=10, color="4A7C59", name="Times New Roman")
    s.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    hdrs = ["Patient Type", "Current Slot (min)", "Recommended (min)",
            "Avg Actual (min)", "Typical Delay (min)", "Delay Rate", "Action"]
    widths = [18, 18, 18, 16, 18, 14, 32]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        _hdr(cell, "2D4A2D")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 24

    rows = [
        ("New Patient",  15, 30, 32, "+17", "77%",  "Increase to 30 min"),
        ("Follow-up",    15, 15, 13, "-2",  "27%",  "No change needed"),
        ("Chronic Case", 15, 45, 45, "+30", "100%", "Increase to 45 min — urgent"),
        ("Emergency",    15, 60, 55, "+40", "100%", "Block 60 min — always overruns"),
        ("Consultation", 15, 20, 19, "+4",  "43%",  "Increase to 20 min"),
    ]
    action_fills = {"No change": ("F0FDF4","15803D"), "Increase": ("FFFBEB","92400E"), "Block": ("FEF2F2","B91C1C")}
    for ri, row in enumerate(rows, 4):
        bg = "F5F8F0" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Times New Roman", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _borders()
            if ci == 2:
                cell.fill = PatternFill("solid", fgColor="FEF2F2")
                cell.font = Font(name="Times New Roman", size=10, color="B91C1C", bold=True)
            elif ci == 3:
                cell.fill = PatternFill("solid", fgColor="F0FDF4")
                cell.font = Font(name="Times New Roman", size=10, color="15803D", bold=True)
            elif ci == 7:
                key = next((k for k in action_fills if str(val).startswith(k)), None)
                if key:
                    cell.fill = PatternFill("solid", fgColor=action_fills[key][0])
                    cell.font = Font(name="Times New Roman", size=10, color=action_fills[key][1])
            else:
                cell.fill = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[ri].height = 24

    ws.merge_cells("A10:G10")
    note = ws["A10"]
    note.value = "Only Follow-up patients fit the 15-min slot. All other types exceed it — especially Chronic Cases and Emergencies (100% delay rate)."
    note.font      = Font(bold=True, size=10, color="2D4A2D", name="Times New Roman")
    note.fill      = PatternFill("solid", fgColor="E8F0E0")
    note.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[10].height = 26


def init_excel():
    if os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Appointments"
    ws.sheet_properties.tabColor = "2D4A2D"

    col_widths = [12, 13, 22, 20, 18, 16, 10, 14, 12, 12, 12, 20]
    for ci, (h, w) in enumerate(zip(HEADERS, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        _hdr(cell)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    _seed(ws)
    _build_slot_guide(wb)
    wb.save(EXCEL_FILE)
    print("Excel created with 100 rows.")


def _next_id():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Appointments"]
    return f"APT{1000 + ws.max_row - 1}"


def book_appointment(data: dict):
    init_excel()
    wb  = openpyxl.load_workbook(EXCEL_FILE)
    ws  = wb["Appointments"]
    row = ws.max_row + 1

    ptype   = data.get("patient_type", "Follow-up")
    alloc   = 15
    actual  = ""
    delay   = ""
    apt_id  = _next_id()

    # Find doctor dept
    dept = next((d["dept"] for d in DOCTORS if d["name"] == data.get("doctor")), "General")

    row_data = [
        apt_id, data["date"], data["patient_name"],
        data["doctor"], dept, ptype,
        data["time_slot"], alloc, actual, delay,
        "Confirmed", data.get("notes", "")
    ]
    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        _data(cell, row)

    wb.save(EXCEL_FILE)
    return {"success": True, "appointment_id": apt_id}


# ── API ───────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    init_excel()
    return render_template("index.html")


@app.route("/api/meta")
def meta():
    return jsonify({"doctors": DOCTORS, "patient_types": PATIENT_TYPES,
                    "time_slots": TIME_SLOTS, "recommended_slots": RECOMMENDED_SLOTS})


@app.route("/api/dashboard")
def dashboard():
    init_excel()
    df = pd.read_excel(EXCEL_FILE, sheet_name="Appointments")
    df = df.fillna("")
    df["Delay_Min"]     = pd.to_numeric(df["Delay_Min"],     errors="coerce").fillna(0)
    df["Actual_Min"]    = pd.to_numeric(df["Actual_Min"],    errors="coerce").fillna(0)
    df["Allocated_Min"] = pd.to_numeric(df["Allocated_Min"], errors="coerce").fillna(15)
    df["Date"]          = pd.to_datetime(df["Date"], errors="coerce")

    total     = len(df)
    confirmed = int((df["Status"] == "Confirmed").sum())
    completed = int((df["Status"] == "Completed").sum())
    cancelled = int((df["Status"] == "Cancelled").sum())
    noshow    = int((df["Status"] == "No-Show").sum())

    # Doctor stats
    doc_stats = []
    for doc, grp in df.groupby("Doctor"):
        dept = grp["Department"].iloc[0] if len(grp) else ""
        comp = grp[grp["Status"] == "Completed"]
        doc_stats.append({
            "doctor":     doc,
            "dept":       dept,
            "total":      len(grp),
            "completed":  len(comp),
            "cancelled":  int((grp["Status"] == "Cancelled").sum()),
            "noshow":     int((grp["Status"] == "No-Show").sum()),
            "confirmed":  int((grp["Status"] == "Confirmed").sum()),
            "comp_rate":  round(len(comp)/len(grp)*100, 1) if len(grp) else 0,
            "avg_delay":  round(comp["Delay_Min"].mean(), 1) if len(comp) else 0,
        })

    # Patient type breakdown
    pt_stats = []
    for pt, grp in df.groupby("Patient_Type"):
        comp = grp[grp["Status"] == "Completed"]
        pt_stats.append({
            "type":    pt,
            "count":   len(grp),
            "avg_actual": round(comp["Actual_Min"].mean(), 1) if len(comp) else 0,
            "recommended": RECOMMENDED_SLOTS.get(pt, 20),
            "avg_delay":   round(comp["Delay_Min"].mean(), 1) if len(comp) else 0,
            "delayed_pct": round((comp["Delay_Min"] > 5).sum() / len(comp) * 100, 1) if len(comp) else 0,
        })

    # Department load
    dept_stats = df.groupby("Department").size().reset_index(name="count")
    dept_list  = [{"dept": r["Department"], "count": int(r["count"])} for _, r in dept_stats.iterrows()]

    # Daily bookings trend (last 14 days)
    df_dated = df.dropna(subset=["Date"]).sort_values("Date")
    daily    = df_dated.groupby(df_dated["Date"].dt.date).size().tail(14)
    trend    = [{"date": str(d), "count": int(v)} for d, v in daily.items()]

    # Status distribution
    status_dist = {"Confirmed": confirmed, "Completed": completed,
                   "Cancelled": cancelled, "No-Show": noshow}

    # Time slot popularity
    slot_counts = df.groupby("Time_Slot").size().sort_values(ascending=False).head(10)
    slots_list  = [{"slot": s, "count": int(c)} for s, c in slot_counts.items()]

    # Patient type by doctor (for stacked chart)
    pt_by_doc = []
    for doc, grp in df.groupby("Doctor"):
        entry = {"doctor": doc.replace("Dr. ", "")}
        for pt in PATIENT_TYPES:
            entry[pt] = int((grp["Patient_Type"] == pt).sum())
        pt_by_doc.append(entry)

    return jsonify({
        "kpis": {"total": total, "confirmed": confirmed, "completed": completed,
                 "cancelled": cancelled, "noshow": noshow,
                 "comp_rate": round(completed/total*100,1) if total else 0},
        "doc_stats":   doc_stats,
        "pt_stats":    pt_stats,
        "dept_stats":  dept_list,
        "trend":       trend,
        "status_dist": status_dist,
        "slots_list":  slots_list,
        "pt_by_doc":   pt_by_doc,
        "recommended_slots": RECOMMENDED_SLOTS,
    })


@app.route("/api/book", methods=["POST"])
def book():
    return jsonify(book_appointment(request.get_json()))


@app.route("/api/appointments")
def appointments():
    init_excel()
    df = pd.read_excel(EXCEL_FILE, sheet_name="Appointments").fillna("")
    df["Date"] = df["Date"].astype(str)
    return jsonify(df.tail(30).to_dict(orient="records"))


@app.route("/api/slots_taken")
def slots_taken():
    """Return booked slots for a given doctor + date."""
    doctor = request.args.get("doctor")
    dt     = request.args.get("date")
    init_excel()
    df = pd.read_excel(EXCEL_FILE, sheet_name="Appointments").fillna("")
    mask = (df["Doctor"] == doctor) & (df["Date"].astype(str) == dt) & (df["Status"] != "Cancelled")
    taken = df[mask]["Time_Slot"].tolist()
    return jsonify({"taken": taken})


@app.route("/api/download")
def download():
    init_excel()
    return send_file(EXCEL_FILE, as_attachment=True, download_name="hospital_appointments.xlsx")


if __name__ == "__main__":
    init_excel()
    app.run(debug=True, port=5000)
